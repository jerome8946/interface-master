# coding:utf-8
import os
from pathlib import Path

import xlrd
from openpyxl import *
from openpyxl.styles import Font, colors

current_path = os.path.abspath('.')


class HandleExcel:
    """封装操作excel的方法"""
    startPath = os.path.join(current_path, 'logs\interface.xlsx')
    startPath = startPath.replace("\\\\", "\\")

    def __init__(self, file=startPath, sheet_id=0):
        my_file = Path(file)
        if my_file.exists() == False:
            # 创建一个excel工作簿，注意该工作簿是在内存中创建
            wb = Workbook()
            # 选择第一个工作表
            ws = wb.create_sheet('Sheet')
            wb.save(file)
        self.file = file
        self.sheet_id = sheet_id
        self.data = self.get_data()
        # 为了在创建一个实例时就获得excel的sheet对象，可以在构造器中调用get_data()
        # 因为类在实例化时就会自动调用构造器，这样在创建一个实例时就会自动获得sheet对象了

    def isExists(self, path):
        result = os.path.exists(path)
        return result

    # 获取某一页sheet对象
    def get_data(self):
        data = xlrd.open_workbook(self.file)
        sheet = data.sheet_by_index(self.sheet_id)
        return sheet

    # 获取excel数据行数
    def get_rows(self):
        rows = self.data.nrows
        # t = self.get_data()  # 调用get_data()取得sheet对象(如果不在构造器获取sheet对象，就需要在方法内先获取sheet对象，再进行下一步操作，每个方法都要这样，所以还是写在构造器中方便)
        # rows = t.nrows
        return rows

    def get_clos(self):
        cols = self.data.ncols

        return cols

    # 获取某个单元格数据
    def get_value(self, row, col):
        value = self.data.cell_value(row, col)
        return value

    # 向某个单元格写入数据
    def write_value(self, path, row, col, value, valuePass=0):
        wb = load_workbook(path)
        ws = wb['Sheet1']
        ws.cell(row=row, column=col).value = value
        if valuePass == 1:
            font = Font(u'宋体', size=10, bold=True, italic=True, color='FF0000')
            ws.cell(row=row, column=col).font = font
        else:
            font = Font(u'宋体', size=10, italic=True, color='000000')
            ws.cell(row=row, column=col).font = font
        wb.save(path)

    def write_value_all(self, path, allList=[]):
        wb = load_workbook(path)
        ws = wb['Sheet1']
        for all in allList:
            ws.cell(row=all['row'], column=all['col']).value = all['value']
            if all['valuePass'] == 1:
                font = Font(u'宋体', size=10, bold=True, italic=True, color='FF0000')
                ws.cell(row=all['row'], column=all['col']).font = font
            else:
                font = Font(u'宋体', size=10, italic=True, color='000000')
                ws.cell(row=all['row'], column=all['col']).font = font
        wb.save(path)


if __name__ == '__main__':
    h = HandleExcel()
    h.write_value('E:\pythonPro\phone建融\phoneAll.xlsx', 4, 1, 'sdsfsdfsdfsdfasdfadff')
